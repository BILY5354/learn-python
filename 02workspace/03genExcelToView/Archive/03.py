import openpyxl
import json
from collections import OrderedDict

EXCELPATH = ".\\02workspace\\03genExcelToView\\Output\\outputExcel.xlsx"
# 现在传的数据一次传一个版本
GETLIST = [
    {'00623-2': {'基准点(REF)': 14, '晶粒位置(DP)': 1, '晶粒污染(DC)': 48,
                 '晶粒爬胶(DG)': 140, '二焊脚起(SL)': 7, '焊点缺失(MB)': 5}},
    {'0424-1': {}},
    {'11111': {'粗铝线焊点宽度(LA2)': 1, '粗铝线焊点高度(LA4)': 7,
               '粗铝线焊点长度(LA7)': 2, '焊线弧高检测错误(MWM)': 5, 'ROI超出图像(ROI)': 1}},
    {'027': {'(Bad app 186 code 0)': 1, '粗铝线焊点位置(LA1)': 5,
             '粗铝线焊点高度(LA4)': 6, '粗线弧高(LH1)': 1, '粗线一焊缺失(LWM)': 3, '基准点(REF)': 2}}
]
# 加载 excel 文件
wb = openpyxl.load_workbook(EXCELPATH)

sheet = wb['Sheet1']


# 获取现 excel 数据集中的缺陷数量
def GetDataSetDefectNu():

    aList = []  # 将第一列全部存入列表中
    for row in range(1, sheet.max_row+1):
        cell = sheet.cell(row=row, column=1)
        aList.append(cell.value)

    result = {}  # 将第一列的缺陷按不同数据集保存到字典中
    rowInfoDict = {}  # 用于保存每个数据集的最后一行行数 方便插入
    rowNu = 1
    current_dataset = None  # 用于保存字典中的 key

    for item in aList:
        if item.startswith('数据集'):
            current_dataset = item.strip()
            result[current_dataset] = {}
        elif item.startswith('修改一下'):
            continue
        elif item.startswith('结'):
            rowInfoDict[current_dataset] = rowNu
        else:  # 缺陷的名称 塞入字典中的 value
            result[current_dataset][item] = 0
        rowNu += 1  # 行数加1
    # print(result)
    return result, rowInfoDict


# * 与原来数据集进行比对
def CompareDSDefect(dsdDict, rowInfoDict, getList):

    toInsDict = {}  # 每个数据集需要新增的缺陷名字
    for i in range(0, len(dsdDict.keys()), 1):
        # 显示获取数据有 左侧列表没有的名称 https://www.cnblogs.com/hightech/p/14164437.html
        diffDefect = list(getList[i].values())[0].keys() - \
            list(dsdDict.values())[i].keys()
        toInsDict[list(dsdDict)[i]] = list(diffDefect)

    print(rowInfoDict)

    newInsertRow = 0
    # todo 插入新行操作: k是数据集x， v是需要插入的缺陷名字
    for k, vList in toInsDict.items():
        # rowInfoDict[k]是最后一行的行数 len(v)是插入的长度
        sheet.insert_rows(rowInfoDict[k] + newInsertRow, len(vList))
        newInsertRow += len(vList)

        # # # 往单元格插入数据
        # i = 0
        # # * rowInfoDict[k]+newInsertRow+1 其中 +1 是代表现在现在需要修改的
        # for row in range(rowInfoDict[k]+newInsertRow+1, rowInfoDict[k]+newInsertRow+len(vList)+1):
        #     sheet.cell(row=row, column=1).value = vList[i]
        #     i += 1  # 下标加1取下一个缺陷名字


dataSetDefect, rowInfo = GetDataSetDefectNu()

CompareDSDefect(dataSetDefect, rowInfo, GETLIST)  # 左侧栏数据 获取到的数据

# dataSetDefect = {1: 0, 2: 0, 3: 0, 4: 0}


# startCol = -1
# # 判断现在有数据的是第几列
# for col in range(1, sheet.max_column+1):
#     cell = sheet.cell(row=1, column=col)
#     if cell.value:
#         print(cell.value)
#     else:
#         startCol = col  # 记录新增的列是更新的位置
#         sheet.insert_cols(col, 1)  # 在没有数据的地方插入 需要插入的列数
#         break

# # 从对应的列第二行开始插入
# for row in range(2, sheet.max_row+1):
#     for col in range(startCol, startCol+1):  # !现在用 startCol+1 代替
#         sheet.cell(row=row, column=col).value = 1

wb.save(EXCELPATH)
