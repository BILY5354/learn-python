#
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import Color

# todo 拿到的数据
GETLIST = [[{'00623-2': {'基准点(REF)': 14, '晶粒位置(DP)': 1, '晶粒污染(DC)': 48, '晶粒爬胶(DG)': 140, '二焊脚起(SL)': 7, '焊点缺失(MB)': 5}}, {'0424-1': {}}, {'11111': {'粗铝线焊点宽度(LA2)': 1, '粗铝线焊点高度(LA4)': 7, '粗铝线焊点长度(LA7)': 2, '焊线弧高检测错误(MWM)': 5, 'ROI超出图像(ROI)': 1}}, {'027': {'(Bad app 186 code 0)': 1, '粗铝线焊点位置(LA1)': 5, '粗铝线焊点高度(LA4)': 6, '粗线弧高(LH1)': 1, '粗线一焊缺失(LWM)': 3, '基准点(REF)': 2}}], [{'23B2326-1-042': {'键合不良(WB)': 8}}, {'23C0117-1-001': {'键合不良(WB)': 3}}, {'23C1442-1-021': {'一焊缺失(BM)': 48, '线追踪(WT)': 1}}, {'23C1442-1-022': {'一焊缺失(BM)': 33}}], [{'23C1442-1-025': {'一焊缺失(BM)': 203}}, {'': {'粗铝线焊点长度(LA7)': 32}}, {'381': {'(Bad app 245 code 0)': 1, '晶粒爬胶(DG)': 2, '粗铝线焊点高度(LA4)': 2, '粗线弧高(LH1)': 6, '基准点(REF)': 1, '焊线变形(WD)': 2}}, {'11111': {'粗铝线焊点高度(LA4)': 3, '基准点(REF)': 1}}], [{'A316047001': {'3D测高缺陷(3D)': 1, '表面异常(DG)': 10, '基准点(REF)': 4}}, {'5184': {'晶粒爬胶(DG)': 19, '焊点缺失(MB)': 4}}, {'518-4': {'晶粒爬胶(DG)': 12, '焊线弧高检测错误(MWM)': 19, '塌线缺陷(SW)': 1}}, {'7-10': {'一焊缺失(BM)': 16, '焊点缺失(MB)': 22, '基准点(REF)': 644, '二焊脚起(SL)': 23, '线追踪(WT)': 6}}], [{'7-11': {'一焊缺失(BM)': 28, '焊点缺失(MB)': 10, '基准点(REF)': 601, '二焊脚起(SL)': 13, '线追踪(WT)': 5}}, {'7-12': {'一焊缺失(BM)': 29, '焊点缺失(MB)': 9, '基准点(REF)': 594, '二焊脚起(SL)': 15, '线追踪(WT)': 5}}, {'7-4': {'一焊缺失(BM)': 22, '焊点缺失(MB)': 14, '基准点(REF)': 639, '二焊脚起(SL)': 18, '线追踪(WT)': 9}}, {'7-5': {'一焊缺失(BM)': 28, '焊点缺失(MB)': 13, '基准点(REF)': 633, '二焊脚起(SL)': 20, '线追踪(WT)': 7}}], [{'7-6': {'一焊缺失(BM)': 17, '焊点缺失(MB)': 15, '基准点(REF)': 658, '二焊脚起(SL)': 13, '线追踪(WT)': 8}}, {'7-7': {'一焊缺失(BM)': 35, '焊点缺失(MB)':10, '基准点(REF)': 610, '二焊脚起(SL)': 6, '线追踪(WT)': 4}}, {'7-8': {'一焊缺失(BM)': 27, '焊点缺失(MB)': 10, '基准点(REF)': 606, '二焊脚起(SL)': 9, '线追踪(WT)': 4}}, {'7-9': {'一焊缺失(BM)': 36, '焊点缺失(MB)': 10, '基准点(REF)': 609, '二焊脚起(SL)': 8, '线追踪(WT)': 4}}], [{'A1': {'焊线弧高检测错误(MWM)': 2, '基准点(REF)': 17}}, {'A1': {'粗铝线焊点位置(LA1)': 5, '粗铝线焊点高度(LA4)': 1, '粗线弧高(LH1)': 5}}, {'A2': {'粗铝线焊点宽度(LA2)': 6, '粗铝线焊点长度(LA7)': 1}}, {'A25000F001.00C': {'晶粒爬胶(DG)': 1, '粗铝线焊点宽度(LA2)': 1, '粗铝线追踪(LA6)': 1, '粗铝线焊点长度(LA7)': 12, '粗线弧高(LH1)': 56, '基准点(REF)': 1}}], [{'A3': {'粗铝线焊点位置(LA1)': 4, '粗铝线焊点高度(LA4)': 1, '粗铝线焊点长度(LA7)': 1, '粗线弧高(LH1)': 5}}, {'A30504602D': {'晶粒爬胶(DG)': 1, '粗铝线焊点高度(LA4)': 2, '粗线弧高(LH1)': 6, '粗线一焊缺失(LWM)':1}}, {'A308018008-jpg': {'(Bad app 245 code 0)': 1, '晶粒爬胶(DG)': 1, '粗铝线焊点高度(LA4)': 1, '塌线(LA9)': 1, '粗线弧高(LH1)': 2}}, {'A316047001': {'3D测高缺陷(3D)': 1, '表面异常(DG)': 10, '基准点(REF)': 4}}], [{'database_2': {'粗铝线焊点高度(LA4)': 16}}, {'EP2305002GMK102': {'晶粒爬胶(DG)': 1, '粗铝线焊点高度(LA4)': 10, '焊点异常(LWM)': 1, '焊线弧高检测错误(MWM)': 1, '基准点(REF)': 2, '塌线缺陷(SW)': 1}}, {'EP2303015DBC074': {'晶粒污染(DC)': 1, '晶粒位置(DP)': 1, '粗铝线焊点长度(LA7)': 5, '粗线弧高(LH1)': 1, '基准点(REF)': 1}}, {'EP2303011GMK078': {'粗铝线焊点位置(LA1)': 1, '粗铝线焊点高度(LA4)': 1, '粗铝线焊点长度(LA7)': 1, '焊点异常(LWM)': 4, '焊线弧高检测错误(MWM)': 5, '塌线缺陷(SW)': 3}}], [{'EP2303006GMK090': {'焊点异常(LA1)': 1, '粗铝线焊点高度(LA4)': 3, '粗铝线焊点长度(LA7)': 1, '粗线弧高(LH1)': 13, '焊点形变或缺失(LWM)': 1, '焊线变形(WD)': 1}}, {'EP2303006GMK014': {'焊点异常(LA1)': 2, '粗铝线焊点高度(LA4)': 3, '粗铝线焊点长度(LA7)': 1, '晶粒爬胶(DG)': 1, '粗线弧高(LH1)': 3, '焊点形变或缺失(LWM)': 3}}, {'EP2303006DBC161': {'粗铝线焊点高度(LA4)': 1, '粗铝线焊点长度(LA7)': 3, '粗线弧高(LH1)': 3, '基准点(REF)': 1, '塌线缺陷(SW)': 3}}, {'EP2303009DBC005': {'晶粒污染(DC)': 2, '粗铝线焊点高度(LA4)': 2, '粗铝线焊点长度(LA7)': 2, '粗线弧高(LH1)': 2, '基准点(REF)': 12}}], [{'grr1': {'基准点(REF)': 1}}, {'grr2': {'晶粒位置(DP)': 1, '测量边缘误差(EME)': 1}}, {'grr3': {'测量边缘误差(EME)': 1}}, {'grr4': {'基准点(REF)': 1}}],[{'grr5': {'基准点(REF)': 1}}, {'grr6': {'基准点(REF)': 1}}, {'grr7': {'基准点(REF)': 1}}, {'grr8': {'基准点(REF)': 1}}]]
# 构造第一行的内容 ver1 ...
titleList = ['']
for i in range(1,len(GETLIST)+1):
    titleList.append(f'ver{i}')

wb = Workbook()
ws = wb.worksheets[0]
ws.append(titleList)  # 添加第一行 ver1...vern


ws = wb.active

# * 设置样式 A-F 列的样式 长与宽
ws.column_dimensions['A'].width = 40
ws.row_dimensions[1].height = 20
aList = ['B', 'C', 'D', 'E', 'F']
for a in aList:
    ws.column_dimensions[a].width = 12

# * 设置填充颜色为浅蓝
blue_color = Color(rgb="00B0F0")
fill = PatternFill(start_color=blue_color,
                   end_color=blue_color, fill_type='solid')

# * 设置浅蓝边框
border_style = Side(style="thin", color='000000')
border = Border(left=border_style, right=border_style,
                top=border_style, bottom=border_style)

# * 设置单元格样式
for x in range(1, 34):
    ws.cell(row=x, column=6).fill = fill
    ws.cell(row=x, column=6).border = border

# * 设置填充颜色为黄色
yellow_color = Color(rgb="FFFF00")
fill = PatternFill(start_color=yellow_color,
                   end_color=yellow_color, fill_type='solid')

border_style = Side(style="thin", color='000000')
border = Border(left=border_style, right=border_style,
                top=border_style, bottom=border_style)

# * 设置单元格样式
for x in range(1, 34):
    for y in range(1,6):
        ws.cell(row=x, column=y).fill = fill
        ws.cell(row=x, column=y).border = border

# 注意相对路径的写法 . 代表的是 LEARN-PYTHON 根目录
wb.save(r'.\\02workspace\\03genExcelToView\\Output\\test1.xlsx')
print("数据存入excel成功")
