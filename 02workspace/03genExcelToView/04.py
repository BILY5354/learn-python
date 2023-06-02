from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles.colors import Color

EXCELPATH = ".\\02workspace\\03genExcelToView\\Output\\test1.xlsx"
# 现在传的数据一次传一个版本
GETDATA = {
    # 一个数据集
    # todo 拿的顺序是 [(11 12) (21 22)] [(33) (44)] [为同一个数据集] (同一个版本中的不同检测项数目)
    "数据集1":
    [
        {
            "id": "20",
            "name": "粗铝线焊点宽度",
            "ver": [{'ver1': 11}, {'ver2': 21}, {'ver3': 5}]  # 每个版本所检测到该缺陷的数量
        },
        {
            "id": "21",
            "name": "粗铝线焊点高度",
            "ver": [{'ver1': 12}, {'ver2': 22}, {'ver3': 6}]
        }
    ],
    # 第二个数据集
    "数据集2":
    [
        {
            "id": "11",
            "name": "测试测试",
            "ver": [{'ver1': 33}, {'ver2': 44}]  # 每个版本所检测到该缺陷的数量
        },
    ],
    "数据集3":
    [
        {
            "id": "234789",
            "name": "测试3",
            # 每个版本所检测到该缺陷的数量
            "ver": [{'ver1': 33}, {'ver2': 34}, {'ver3': 35}, {'ver4': 36}]
        },
        {
            "id": "234790",
            "name": "测试4",
            # 每个版本所检测到该缺陷的数量
            "ver": [{'ver1': 44}, {'ver2': 45}, {'ver3': 46}, {'ver4': 47}]
        },
        {
            "id": "234791",
            "name": "测试5",
            # 每个版本所检测到该缺陷的数量
            "ver": [{'ver1': 55}, {'ver2': 56}, {'ver3': 57}, {'ver4': 58}]
        },
    ],
}

GETYieldTime = {
    'EP2303003GMK125':
    [
        ['20230522', '89.47%', '2分15秒'], ['20230525', '89.47%', '2分25秒'], ['20230525-1',
                                                                           '89.47%', '2分24秒'], ['20230529', '89.47%', '2分24秒'], ['20230529-T8', '89.47%', '2分27秒']
    ],
    'EP2303003GMK126':
    [
        ['20230522', '100.00%', '1分8秒'], ['20230525', '100.00%', '1分9秒'], ['20230525-1',
                                                                           '100.00%', '1分5秒'], ['20230529', '100.00%', '1分8秒'], ['20230529-T8', '100.00%', '1分9秒']
    ],
    'EP2304005GMK042':
    [
        ['20230525', '42.11%', '11分56秒'], ['20230525-1', '42.11%', '11分51秒'], ['20230525-1',
                                                                               '42.11%', '11分38秒'], ['20230529', '42.11%', '11分21秒'], ['20230529-T8', '42.11%', '10分56秒']
    ],
    'EP2304005GMK048':
    [
        ['20230525', '50.00%', '13分10秒'], ['20230525-1', '50.00%',
                                           '12分30秒'], ['20230529', '50.00%', '11分48秒'], ['20230529-T8', '50.00%', '11分34秒']
    ],
    'EP2302093GMK115':
    [
        ['20230525', '15.79%', '11分28秒'], ['20230525-1', '15.79%',
                                           '11分26秒'], ['20230529', '15.79%', '10分48秒'], ['20230529-T8', '15.79%', '10分47秒']
    ],
    'EP2302093GMK141':
    [
        ['20230525', '84.21%', '0分59秒'], ['20230525-1', '84.21%',
                                          '0分57秒'], ['20230529', '84.21%', '0分55秒'], ['20230529-T8', '84.21%', '0分57秒']
    ],
    'EP2302093GMK142':
    [
        ['20230525', '88.89%', '0分56秒'], ['20230525-1', '88.89%',
                                          '0分54秒'], ['20230529', '88.89%', '0 分52秒'], ['20230529-T8', '88.89%', '0分55秒']
    ], 'EP2303009GMK028': [['20230525', '0.00%', '12 分6秒'], ['20230525-1', '0.00%', '11分51秒'], ['20230529', '0.00%', '11分9秒'], ['20230529-T8', '0.00%', '11分48秒']], 'EP2303009GMK125': [['20230525', '10.00%', '12分14秒'], ['20230525-1', '10.00%', '11分51秒'], ['20230529', '10.00%', '11分9秒'], ['20230529-T8', '10.00%', '11分56秒']], 'EP2303016GMK020': [['20230525', '5.26%', '11分36秒'], ['20230525-1', '5.26%', '11分16秒'], ['20230529', '5.26%', '10分29秒'], ['20230529-T8', '5.26%', '11分1秒']], 'EP2303011GMK080': [['20230525', '0.00%', '12分1秒'], ['20230525-1', '0.00%', '11分54秒'], ['20230529', '0.00%', '10 分56秒'], ['20230529-T8', '0.00%', '11分40秒']]}


green = Color(rgb="00B050")
green_fill = PatternFill(start_color=green,
                         end_color=green, fill_type='solid')

yellow = Color(rgb="FFFF00")
yellow_fill = PatternFill(start_color=yellow,
                          end_color=yellow, fill_type='solid')

blue = Color(rgb="00B0F0")
blue_fill = PatternFill(start_color=blue,
                        end_color=blue, fill_type='solid')

border_style = Side(style="thin", color='000000')
black_border = Border(left=border_style, right=border_style,
                      top=border_style, bottom=border_style)


# 将字典所有键存入列表中 存的是数据集名称
getDKEY = [i for i in GETDATA.keys()]
yieldTimeKey = [i for i in GETYieldTime.keys()]

wb = Workbook()
DATASETNU = len(GETDATA)  # 数据集数量
DATAYIELDTIMENU = len(GETYieldTime)


ws = wb.active
ws.title = "汇总表格"  # 将第一个改名
# 有几个数据集 创建多少个
for i in range(0, DATASETNU, 1):
    wb.create_sheet(f'{getDKEY[i]}')


# # * 塞入数据 注意 currentDSNu 下标从 0 开始
# for currentDSNu in range(0, DATASETNU, 1):
#     ws = wb[f'{getDKEY[currentDSNu]}']  # 选中对应的工作簿
#     oneDataSet = GETDATA[getDKEY[currentDSNu]]  # 一个数据集列表
#     oneDataSetVer = oneDataSet[0]['ver']  # 一个数据集对应版本数列表

#     # 开始插入缺陷名称 注意从第二列开始插入
#     for col in range(0, len(oneDataSet), 1):
#         editCol = int(2 + col)
#         ws.cell(
#             row=1, column=editCol).value = oneDataSet[col]['name']
#         ws.cell(row=1, column=editCol).fill = green_fill
#         ws.cell(row=1, column=editCol).border = black_border

#     # 插入 第一列数据集版本号
#     verNu = 1
#     for row in range(0, len(oneDataSetVer), 1):
#         editRow = int(2+row)  # 从第二行开始操作
#         ws.cell(row=editRow,
#                 column=1).value = f'{list(oneDataSetVer[row])[0]}'
#         ws.cell(row=editRow, column=1).fill = yellow_fill
#         ws.cell(row=editRow, column=1).border = black_border
#         verNu += 1

#     # 开始插入数据 按行插入
#     verNu = 0  # 对应版本号 从0开始
#     # todo GETDATA[currentDSNu][0]['ver'] 中的 0 固定死因为每个括号包含的版本数是一样的
#     for row in range(0, len(oneDataSetVer), 1):
#         for col in range(0, len(oneDataSet), 1):
#             # 特定数据集版本中的字典(数量为1)
#             specificDataSetVer = oneDataSet[col]['ver'][verNu]
#             editRow = int(2+row)  # 行数从第二行开始操作
#             editCol = int(2+col)  # 列从第二列开始操作

#             # 获取对应数据集缺陷版本的数量
#             specificKey = [i for i in specificDataSetVer.keys()][0]

#             ws.cell(row=editRow,
#                     column=editCol).value = specificDataSetVer[specificKey]
#             # print(specificDataSetVer)

#         verNu += 1

#     # 开始画折线图
#     lineChart = LineChart()
#     chartData = Reference(ws, min_row=1, max_row=ws.max_row,
#                           min_col=2, max_col=ws.max_column)
#     titleData = Reference(ws, min_row=2, max_row=ws.max_row,
#                           min_col=1, max_col=1)

#     series = Series(chartData)
#     series.marker

#     # todo from_rows=False 默认以没一列为数据系列
#     lineChart.title = f'{getDKEY[currentDSNu]}报告'
#     lineChart.add_data(chartData, from_rows=False, titles_from_data=True)
#     lineChart.set_categories(titleData)
#     ws.add_chart(lineChart, f'F{ws.max_row+10}')


# * 插入汇总表格数据
ws = wb["汇总表格"]

ws.cell(row=1, column=1).value = "良率"
ws.cell(row=1, column=1).fill = blue_fill
ws.cell(row=1, column=1).border = black_border

# 开始插入批次信息 从第二列开始插入
for col in range(0, DATAYIELDTIMENU, 1):
    #! 修改下超链接
    aa = "数据集1"
    link = "test1.xlsx"+f"#{aa}!A1"
    editCol = int(2 + col)
    ws.cell(
        row=1, column=editCol).value = yieldTimeKey[col]
    ws.cell(row=1, column=editCol).style = "Hyperlink"
    ws.cell(row=1, column=editCol).fill = green_fill
    ws.cell(row=1, column=editCol).border = black_border
    ws.cell(row=1, column=editCol).hyperlink = link

#! 现在假设用第一个批次号作版本插入的列表  后面需要传入所有版本的列表
DATAVER = [i[0] for i in GETYieldTime[yieldTimeKey[0]]]  # 存入所有版本号

# 开始插入每个版本信息 从二行开始插入
for row in range(0, len(DATAVER), 1):
    editRow = int(2+row)  # 从第二行开始操作
    ws.cell(row=editRow,
            column=1).value = DATAVER[row]
    ws.cell(row=editRow, column=1).fill = yellow_fill
    ws.cell(row=editRow, column=1).border = black_border

# 开始插入每个批次对应的单元格信息
for col in range(0, DATAYIELDTIMENU, 1):
    specificYieldTime = GETYieldTime[yieldTimeKey[col]]

    insertNu = 0  # 单个数据集中 插入的良率位置 从0开始
    for row in range(0, len(DATAVER), 1):
        editRow = int(2+row)
        editCol = int(2+col)

        # 判断该良率版本是否列相同
        if DATAVER[row] == specificYieldTime[insertNu][0]:
            # print(f'row{editRow} col{editCol} {specificYieldTime[insertNu][0]}')
            value = float(specificYieldTime[insertNu][1].strip('%')) / 100
            ws.cell(row=editRow, column=editCol).value = value
            # #! 增加超链接功能
            # ws.cell(row=editRow, column=editCol).style = "Hyperlink"
            # # link = fr''
            # # ws.cell(row=editRow,column=editCol).hyperlink = link
            insertNu += 1
        else:
            # print(f'row{editRow} col{editCol}')
            continue

# 开始画折线图
lineChart = LineChart()
chartData = Reference(ws, min_row=1, max_row=ws.max_row,
                      min_col=2, max_col=ws.max_column)
titleData = Reference(ws, min_row=2, max_row=ws.max_row,
                      min_col=1, max_col=1)

series = Series(chartData)
series.marker

# todo from_rows=False 默认以没一列为数据系列
lineChart.title = f'汇总良率报告 单位(%)'
lineChart.add_data(chartData, from_rows=False, titles_from_data=True)
lineChart.set_categories(titleData)
lineChart.y_axis.scaling.max = 1 # 设置Y轴  最大值
ws.add_chart(lineChart, f'F{ws.max_column+6}')


# * 第二部分
BottomRow = ws.max_row+2
ws.cell(row=BottomRow, column=1).value = "时间"
ws.cell(row=BottomRow, column=1).fill = blue_fill
ws.cell(row=BottomRow, column=1).border = black_border

# 开始插入批次信息 从第二列开始插入
for col in range(0, DATAYIELDTIMENU, 1):
    #! 修改下超链接
    aa = "数据集1"
    link = "test1.xlsx"+f"#{aa}!A1"
    editCol = int(2 + col)
    ws.cell(
        row=BottomRow, column=editCol).value = yieldTimeKey[col]
    ws.cell(row=BottomRow, column=editCol).style = "Hyperlink"
    ws.cell(row=BottomRow, column=editCol).fill = green_fill
    ws.cell(row=BottomRow, column=editCol).border = black_border
    ws.cell(row=BottomRow, column=editCol).hyperlink = link

# 开始插入每个版本信息 从二行开始插入
for row in range(0, len(DATAVER), 1):
    editRow = int(BottomRow+row+1)  # 从底部加1行插入
    link = "test1.xlsx"+f"#{aa}!A1"
    ws.cell(row=editRow,
            column=1).value = DATAVER[row]
    ws.cell(row=editRow, column=1).fill = yellow_fill
    ws.cell(row=editRow, column=1).border = black_border
    ws.cell(row=editRow, column=1).style = "Hyperlink"
    ws.cell(row=editRow, column=1).hyperlink = link



# 开始插入每个批次对应的单元格信息
for col in range(0, DATAYIELDTIMENU, 1):
    specificYieldTime = GETYieldTime[yieldTimeKey[col]]

    insertNu = 0  # 单个数据集中 插入的良率位置 从0开始
    for row in range(0, len(DATAVER), 1):
        editRow = int(BottomRow+row+1)
        editCol = int(2+col)

        # 判断该良率版本是否列相同
        if DATAVER[row] == specificYieldTime[insertNu][0]:
            minutes, seconds = specificYieldTime[insertNu][2].split("分")
            seconds = seconds.rstrip("秒")
            time_decimal = float(minutes) + float(seconds) / 60
            value = "{:.2f}".format(time_decimal)

            ws.cell(row=editRow,
                    column=editCol).value = float(value)
            #! 增加超链接功能
            # ws.cell(row=editRow, column=editCol).style = "Hyperlink"
            # link = fr''
            # ws.cell(row=editRow,column=editCol).hyperlink = link
            insertNu += 1
        else:
            continue

# 开始画折线图
lineChart2 = LineChart()
chartData2 = Reference(ws, min_row=BottomRow, max_row=ws.max_row,
                      min_col=2, max_col=ws.max_column)
titleData2 = Reference(ws, min_row=BottomRow, max_row=ws.max_row,
                      min_col=1, max_col=1)

series2 = Series(chartData2)
series2.marker

# todo from_rows=False 默认以没一列为数据系列
lineChart2.title = f'汇总时间报告 单位时间'
lineChart2.add_data(chartData2, from_rows=False, titles_from_data=True)
lineChart2.set_categories(titleData2)
ws.add_chart(lineChart2, f'F{ws.max_column+20}')

wb.save(EXCELPATH)
