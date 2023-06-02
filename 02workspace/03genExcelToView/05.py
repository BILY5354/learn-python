from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

EXCELPATH = ".\\02workspace\\03genExcelToView\\Output\\test2.xlsx"

#* 测试超链接
wb = Workbook()
wb.create_sheet('ver2')
ws = wb.active
ws = wb["Sheet"]
ws.cell(row=1,column=1).value = "ver2"
ws.cell(row=1,column=1).style="Hyperlink"
link = "test2.xlsx"+"#ver2!A1"
ws.cell(row=1,column=1).hyperlink = link


ws.cell(row=1,column=2).value = "ver2"
ws.cell(row=1,column=2).style="Hyperlink"
link = r"E:\learn-python\learn-python\02workspace\03genExcelToView\Output\EP2304005GMK042_20230525_021601.db3"
ws.cell(row=1,column=2).hyperlink = link


# # 创建超链接对象并添加到单元格
# link = "#ver2!A1"
# ws.cell(row=1, column=1).value = 'Link to another sheet'
# ws.cell(row=1, column=1).hyperlink = (link)

# 在单元格中存储文件路径
# ws.cell(row=1, column=1).value = 'Link to local file'
# ws.cell(row=1, column=1).hyperlink = r'\\path\to\file'

wb.save(EXCELPATH)